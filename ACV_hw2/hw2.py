import numpy as np
import cv2
from math import pi
import openpyxl
import matplotlib.pyplot as plt

start_point = [[2252, 1738], [2324, 1653], [2317, 1620], [2220, 1859], [2362, 1757], [2382, 1398], [2115, 1980],
               [2427, 1820], [2269, 1308]]
image_size = [4672, 3104]
block_size = [50, 150, 250]
search_range18 = [15, 50, 80, 120]   # search radium
search_range53 = [30, 100, 150, 300]   # search radium
search_range135 = [100, 200, 300, 500]   # search radium
sensor_width = 23.4
f = [18, 53, 135]
mm_moved = [0, 1, 5, 10, 20]
mm_moved1 = [1, 5, 10, 20]
object_distance = [600, 1200, 1800]


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


book_name_xlsx = 'data.xlsx'

sheet_name_xlsx = 'data1'


#file_path1 = './Photo/18mm/600mm_0mm.jpg'  # Path of JPG file
#file_path2 = './Photo/18mm/600mm_1mm.jpg'  # Path of JPG file
file_path_18_600 = []
file_path_18_1200 = []
file_path_18_1800 = []

file_path_53_600 = []
file_path_53_1200 = []
file_path_53_1800 = []

file_path_135_600 = []
file_path_135_1200 = []
file_path_135_1800 = []

image18_600 = []
image18_1200 = []
image18_1800 = []

image53_600 = []
image53_1200 = []
image53_1800 = []

image135_600 = []
image135_1200 = []
image135_1800 = []


for i in range(5):
    file_path_18_600.append('./Photo/18mm/600mm_' + str(mm_moved[i]) +'mm.jpg')
    file_path_18_1200.append('./Photo/18mm/1200mm_' + str(mm_moved[i]) +'mm.jpg')
    file_path_18_1800.append('./Photo/18mm/1800mm_' + str(mm_moved[i]) + 'mm.jpg')
    file_path_53_600.append('./Photo/53mm/600mm_' + str(mm_moved[i]) +'mm.jpg')
    file_path_53_1200.append('./Photo/53mm/1200mm_' + str(mm_moved[i]) +'mm.jpg')
    file_path_53_1800.append('./Photo/53mm/1800mm_' + str(mm_moved[i]) + 'mm.jpg')
    file_path_135_600.append('./Photo/135mm/600mm_' + str(mm_moved[i]) +'mm.jpg')
    file_path_135_1200.append('./Photo/135mm/1200mm_' + str(mm_moved[i]) +'mm.jpg')
    file_path_135_1800.append('./Photo/135mm/1800mm_' + str(mm_moved[i]) + 'mm.jpg')

for i in range(5):
    image18_600.append(cv2.imread(file_path_18_600[i]))
    image18_1200.append(cv2.imread(file_path_18_1200[i]))
    image18_1800.append(cv2.imread(file_path_18_1800[i]))
    image53_600.append(cv2.imread(file_path_53_600[i]))
    image53_1200.append(cv2.imread(file_path_53_1200[i]))
    image53_1800.append(cv2.imread(file_path_53_1800[i]))
    image135_600.append(cv2.imread(file_path_135_600[i]))
    image135_1200.append(cv2.imread(file_path_135_1200[i]))
    image135_1800.append(cv2.imread(file_path_135_1800[i]))




# Block Matching Function
def block_matching(trucka, truckb, block_size, search_range, start_point):
    # trucka: template(T), truckb: changed image(I)
    x = int(start_point[0] - block_size / 2)
    y = int(start_point[1] - block_size / 2)
    width = image_size[0]
    height = image_size[1]

    best_sad = float('inf')     # 正無窮
    best_dx, best_dy = 0, 0

    # Define the search range for the current block
    search_positions = range(-search_range, search_range + 1)
    for dy in search_positions:
        for dx in search_positions:
            y1, x1 = y + dy, x + dx
            if 0 <= x1 < width - block_size and 0 <= y1 < height - block_size:
                block_a = trucka[y:y + block_size, x:x + block_size]
                block_b = truckb[y1:y1 + block_size, x1:x1 + block_size]
                # Sum of Absolute Differences (SAD)
                # sad = np.sum(np.abs(block_a - block_b))
                sad = np.sum(np.square(block_a-block_b))
                if sad < best_sad:
                    best_sad = sad
                    best_dx, best_dy = dx, dy

    motion_vectors = [best_dx, best_dy]

    return motion_vectors

# Perform block matching for each block size
motion_vectors = []
for i in range(4):
    motion_vectors.append(block_matching(image18_600[0], image18_600[i+1], block_size[0], search_range18[i], start_point[0]))
for i in range(4):
    motion_vectors.append(block_matching(image18_1200[0], image18_1200[i + 1], block_size[0], search_range18[i], start_point[1]))
for i in range(4):
    motion_vectors.append(block_matching(image18_1800[0], image18_1800[i + 1], block_size[0], search_range18[i], start_point[2]))
for i in range(4):
    motion_vectors.append(block_matching(image53_600[0], image53_600[i + 1], block_size[1], search_range53[i], start_point[3]))
for i in range(4):
    motion_vectors.append(block_matching(image53_1200[0], image53_1200[i + 1], block_size[1], search_range53[i], start_point[4]))
for i in range(4):
    motion_vectors.append(block_matching(image53_1800[0], image53_1800[i + 1], block_size[1], search_range53[i], start_point[5]))
for i in range(4):
    motion_vectors.append(block_matching(image135_600[0], image135_600[i + 1], block_size[2], search_range135[i], start_point[6]))
for i in range(4):
    motion_vectors.append(block_matching(image135_1200[0], image135_1200[i + 1], block_size[2], search_range135[i], start_point[7]))
for i in range(4):
    motion_vectors.append(block_matching(image135_1800[0], image135_1800[i + 1], block_size[2], search_range135[i], start_point[8]))

# print(f"Motion Vectors for block size {block_size}x{block_size}:")
print(motion_vectors)


# Visualization function
def draw_motion_vectors(image, motion_vectors, start_point):
    # Starting point of the vector
    start_y = start_point[1]
    start_x = start_point[0]

    # The motion vector
    dy = motion_vectors[1]
    dx = motion_vectors[0]

    # Draw the motion vector if it is not zero
    if dy != 0 or dx != 0:
        image_with_arrow = cv2.arrowedLine(image, (start_x, start_y), (dx + start_x, dy + start_y),(225, 225, 225), 2)


    # Show the plot
    #plt.imshow(image_with_arrow)
    #plt.axis('off')
    #plt.show()
    cv2.namedWindow('image', cv2.WINDOW_NORMAL)
    cv2.imshow('image', image_with_arrow)
    cv2.waitKey(0)

# Call the function with a block size of 9 (replace this with your real data)
for i in range(4):
    draw_motion_vectors(image18_600[i+1], motion_vectors[i], start_point[0])
for i in range(4):
    draw_motion_vectors(image18_1200[i+1], motion_vectors[i+4], start_point[1])
for i in range(4):
    draw_motion_vectors(image18_1800[i+1], motion_vectors[i+8], start_point[2])
for i in range(4):
    draw_motion_vectors(image53_600[i+1], motion_vectors[i+12], start_point[3])
for i in range(4):
    draw_motion_vectors(image53_1200[i+1], motion_vectors[i+16], start_point[4])
for i in range(4):
    draw_motion_vectors(image53_1800[i+1], motion_vectors[i+20], start_point[5])
for i in range(4):
    draw_motion_vectors(image135_600[i+1], motion_vectors[i+24], start_point[6])
for i in range(4):
    draw_motion_vectors(image135_1200[i+1], motion_vectors[i+28], start_point[7])
for i in range(4):
    draw_motion_vectors(image135_1800[i+1], motion_vectors[i+32], start_point[8])
def FOV_theoretical():
    FOV_the = []
    for i in range(3):
        FOV_the.append(2 * np.arctan(sensor_width/(2*f[i]))*180/pi)
    return FOV_the
def FOV_measured():
    FOV_value = []
    for j in range(3):
        for i in range(4):
            FOV_value.append(2 * np.arctan((image_size[0] * (float(mm_moved[i+1]) / motion_vectors[i+j*12][0])) / (2*object_distance[0]))*180/ pi)
            FOV_value.append(2 * np.arctan(
                (image_size[0] * (float(mm_moved[i + 1]) / motion_vectors[i+4+j*12][0])) / (2 * object_distance[1])) * 180 / pi)
            FOV_value.append(2 * np.arctan(
                (image_size[0] * (float(mm_moved[i + 1]) / motion_vectors[i + 8+j*12][0])) / (2 * object_distance[2])) * 180 / pi)
    return FOV_value
fov_theoretical = FOV_theoretical()
fov_measured = FOV_measured()
print('FOV_theoretical=', fov_theoretical)
print('FOV_measured=', fov_measured)
# print(2*np.arctan((image_size[0] * (mm_moved[2+1] / 52)) / (2*object_distance[0]))*180/ pi)

excel_data = [["相機焦距(mm)", "物體距離(mm)", "物體實際位移(mm)", "物體位移(pixels)", "mm/pixel", "FOV估計值", "FOV理論值"]]
for i in range(3):
    for j in range(3):
        for k in range(4):
            excel_data.append([str(f[i])])
            excel_data[i*12+j*4+k+1].append(str(object_distance[j]))
            excel_data[i*12+j*4+k+1].append(str(mm_moved1[k]))
            excel_data[i * 12 + j * 4 + k + 1].append(str(motion_vectors[i * 12 + j * 4 + k][0]))
            excel_data[i*12+j*4+k+1].append(str(round(float(mm_moved1[k])/motion_vectors[i*12+j*4+k][0], 6)))
            excel_data[i*12+j*4+k+1].append(str(round(fov_measured[i*12+j*4+k], 6)))
            excel_data[i*12+j*4+k+1].append(str(round(fov_theoretical[i], 6)))
write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, excel_data)